"""
Build blog posts for sfsmodels.org. Each post targets one long-tail SEO keyword.

Run:
    cd financial-models-website && python3 build_blog_posts.py

Output: HTML files in /blog/.
"""

from pathlib import Path

POSTS = [
    {
        "slug": "mid-year-discounting-dcf",
        "title": "How to Calculate Mid-Year Discounting in a DCF Model",
        "h1": "How to Calculate Mid-Year Discounting in a DCF Model",
        "keyword": "mid year discounting dcf",
        "meta_desc": "Mid-year discounting in DCF models corrects the systematic ~5% under-valuation from end-of-year discounting. Formulas, Excel implementation, common mistakes.",
        "published": "2026-04-21",
        "summary": "End-of-year discounting under-discounts cash flows by half a year of WACC. Mid-year discounting fixes this. Here's the formula, the Excel implementation, and the common mistakes.",
        "product_link": "/dcf-model-excel.html",
        "product_name": "DCF Model Excel Template",
        "body": """<p>If you've built a DCF model in Excel and used standard end-of-year discounting, you're systematically under-valuing the business by approximately half a year of WACC. On a growth business, that's a 5-10% understatement of enterprise value. Mid-year discounting fixes this.</p>

<h2>What is mid-year discounting?</h2>

<p>In a standard DCF, you discount each year's free cash flow (FCF) back to present value using:</p>

<pre><code>PV = FCF / (1 + WACC)^t</code></pre>

<p>Where <code>t</code> is the number of years from the valuation date.</p>

<p>The problem: this formula assumes all of year <code>t</code>'s cash flow arrives on December 31st of year <code>t</code>. In reality, cash flows arrive throughout the year &mdash; roughly evenly. The "average" cash arrives at mid-year (June 30th).</p>

<p>Mid-year discounting adjusts for this by discounting cash flows by half a period less:</p>

<pre><code>PV = FCF / (1 + WACC)^(t - 0.5)</code></pre>

<p>For a 5-year forecast at 10% WACC, mid-year discounting gives you ~4.7% higher PV vs end-of-year. On a $1bn EV business, that's $47m of value.</p>

<h2>When to use mid-year vs end-of-year</h2>

<p><strong>Use mid-year when:</strong></p>
<ul>
<li>Cash flows arrive evenly throughout the year (most operating businesses)</li>
<li>You're presenting valuations to investment committees who expect mid-year (standard at most US PE firms)</li>
<li>Public company valuations (most sell-side equity research uses mid-year)</li>
</ul>

<p><strong>Use end-of-year when:</strong></p>
<ul>
<li>Cash flows are heavily back-loaded (project finance, infrastructure concessions where cash arrives at completion milestones)</li>
<li>You're modelling for tax-driven structures where year-end timing matters</li>
<li>The receiving institution has a stated convention (some EU banks default to end-of-year)</li>
</ul>

<h2>The terminal value adjustment</h2>

<p>Most DCF builders forget that mid-year discounting also affects terminal value. Two adjustments:</p>

<p><strong>1. Discount the terminal value back further.</strong></p>

<p>If your explicit forecast ends in year 5 and you've used mid-year discounting for years 1&ndash;5, the terminal value (which represents value FROM year 6 onwards) should be discounted at:</p>

<pre><code>TV PV = TV / (1 + WACC)^(5 - 0.5) = TV / (1 + WACC)^4.5</code></pre>

<p>NOT <code>TV / (1 + WACC)^5</code>.</p>

<p><strong>2. The terminal value formula doesn't change.</strong></p>

<p>The Gordon growth formula gives you the value AS OF the end of the explicit forecast period. So whether you use mid-year or end-of-year, the TV formula itself is:</p>

<pre><code>TV = FCF(year n+1) / (WACC - g)</code></pre>

<p>Only the discount factor applied to bring TV back to present value changes.</p>

<h2>Implementing mid-year discounting in Excel</h2>

<p>Add a "discount factor" row to your DCF tab that applies mid-year automatically:</p>

<pre><code>Year:           1     2     3     4     5
FCF:            100   120   140   160   180
Discount factor: 1/(1+WACC)^0.5  1/(1+WACC)^1.5  1/(1+WACC)^2.5  ...

PV = FCF &times; Discount factor</code></pre>

<p>Sum the PVs of years 1&ndash;n + discounted terminal value = enterprise value.</p>

<p>For our <a href="/dcf-model-excel.html">DCF Model</a>, we built in a toggle on the INPUTS tab &mdash; switch between mid-year and end-of-year and the entire model recalibrates.</p>

<h2>Common mistakes</h2>

<ol>
<li><strong>Discounting terminal value at year n + 1 instead of n.</strong> TV represents value at end of year n, so discount it at year n (or year n &minus; 0.5 for mid-year).</li>
<li><strong>Using mid-year for one year and end-of-year for others.</strong> Be consistent across all years in the explicit forecast.</li>
<li><strong>Ignoring stub periods.</strong> If your valuation date isn't year-end, the first period is a stub (e.g., 7 months). Adjust the discount factor for the stub period accordingly.</li>
<li><strong>Mid-year on a project that ends in year 5.</strong> If the business has a defined termination (project finance), don't use mid-year for the final year &mdash; use the actual cash receipt timing.</li>
</ol>

<h2>Free DCF model</h2>

<p>If you want to skip the build, our <a href="/free-sample.html">free DCF Lite</a> has both end-of-year and mid-year discounting available as a switch.</p>

<p>For the full version with three-statement integration, terminal value cross-check, and 8 integrity checks: <a href="/dcf-model-excel.html">DCF Model Excel Template</a>.</p>
""",
    },
    {
        "slug": "deposit-beta-calibration",
        "title": "Deposit Beta in Bank ALM Models: 2022-2025 Calibration",
        "h1": "Deposit Beta in Bank ALM Models: The 2022-2025 Calibration Problem",
        "keyword": "deposit beta calibration",
        "meta_desc": "Deposit beta assumptions calibrated against 2014-2019 are wrong by ~3x for the post-2022 cycle. How to recalibrate by deposit tier and reflect the SVB lessons.",
        "published": "2026-04-21",
        "summary": "Most US bank ALM models use deposit beta from the 2014-2019 cycle. Actual betas in 2022-2025 ran 2-3x higher. The recalibration framework, by tier, with EVE/NII shock implications.",
        "product_link": "/bank-financial-model.html",
        "product_name": "Bank Long-Term Plan Model",
        "body": """<p>Most US bank ALM models still use deposit beta assumptions calibrated against the 2014&ndash;2019 rate cycle. The 2022&ndash;2025 cycle showed those calibrations were materially wrong. Banks running stale betas understated their NII risk by ~3x heading into the SVB crisis.</p>

<h2>What is deposit beta?</h2>

<p>Deposit beta is the percentage of a base rate (typically Fed Funds for US, Bank Rate for UK) that gets passed through to the rate paid on a deposit product.</p>

<p>If Fed Funds rises 100bps and a savings account rate rises 30bps, the deposit beta is 30%.</p>

<p>Deposit beta varies dramatically by:</p>
<ul>
<li><strong>Product type:</strong> instant access &lt; notice &lt; fixed term</li>
<li><strong>Customer type:</strong> retail &lt; small business &lt; commercial &lt; institutional</li>
<li><strong>Insurance status:</strong> insured (FDIC-covered) &lt; uninsured</li>
<li><strong>Channel:</strong> branch &lt; online &lt; broker</li>
</ul>

<h2>What 2022-2025 changed</h2>

<p>The previous decade (2010&ndash;2021) was characterised by ZIRP and modest rate moves. Deposit betas calibrated against this period showed:</p>

<ul>
<li>Retail savings: 20&ndash;30%</li>
<li>Commercial deposits: 30&ndash;50%</li>
<li>Institutional / brokered: 60&ndash;80%</li>
</ul>

<p>When rates rose 525bps over 2022&ndash;2023, actual betas were materially higher:</p>

<ul>
<li>Retail savings: 35&ndash;50% (up from 20&ndash;30%)</li>
<li>Commercial deposits: 70&ndash;90% (up from 30&ndash;50%)</li>
<li>Uninsured commercial: 80&ndash;95% (the SVB cohort)</li>
</ul>

<p>The reason: depositors had access to better rate alternatives (T-bills, money market funds yielding 5%+) and digital channels made it easy to move money.</p>

<h2>Why this matters for ALM models</h2>

<p>ALM models project NII under interest rate scenarios. Two key levers:</p>

<ol>
<li>Asset side: how fast do loan yields reprice</li>
<li>Liability side: how fast do deposit rates reprice (driven by deposit beta)</li>
</ol>

<p>If your model assumes 30% beta when actual beta is 80%, in a +200bps shock your modelled NII is materially overstated. Your EVE / NII shock numbers are similarly distorted.</p>

<p>For the 2022&ndash;23 cycle, banks running 2019-vintage betas were modelling NII benefits of ~10&ndash;15%. Actual NII benefits were closer to 3&ndash;5% &mdash; and for some banks (those with high uninsured commercial mix), NII actually compressed.</p>

<h2>How to recalibrate</h2>

<p><strong>1. Segment deposits by beta tier.</strong></p>

<p>Don't average across the whole book. Split into:</p>
<ul>
<li>Insured retail (low beta)</li>
<li>Insured commercial (medium beta)</li>
<li>Uninsured commercial (high beta)</li>
<li>Brokered / wholesale (highest beta)</li>
</ul>
<p>Each gets its own beta assumption.</p>

<p><strong>2. Use 2022-2025 actuals as base case.</strong></p>

<p>Calibrate against your bank's actual experience over the rising rate cycle. Don't blend with pre-2022 data &mdash; the rate environment was structurally different (digital access, money market yields, etc.).</p>

<p><strong>3. Flex beta in scenarios.</strong></p>

<p>Base case = 2022&ndash;2025 actuals. Adverse scenario = 1.2x base (faster pass-through under stress). Severely adverse = 1.5x base (run-off scenario).</p>

<p><strong>4. Re-calculate EVE / NII shock.</strong></p>

<p>Re-run your standard shock scenarios (+200bps parallel, etc.) with new betas. Expect your numbers to look materially worse than the previous cycle's output. That's the point &mdash; the previous cycle's output was wrong.</p>

<h2>What about the deposit beta floor?</h2>

<p>Some banks model a "floor" beta below which the deposit rate doesn't fall in a falling rate environment. For example, retail savings rates might have a 0.10% floor regardless of how low Fed Funds goes.</p>

<p>The floor matters in falling-rate scenarios. If you model 30% beta on the way up but ignore the floor on the way down, your NII benefit in a rate cut scenario is overstated.</p>

<p>For a complete bank long-term plan model that handles deposit beta tiering, EVE / NII shock, and floor logic, see our <a href="/bank-financial-model.html">Bank Long-Term Plan Model</a>.</p>

<h2>What to do if you can't recalibrate immediately</h2>

<p>If your ALM model is stuck on old betas and a recalibration is months away, the interim hack: apply a "beta haircut" to your modelled NII benefit.</p>

<p>For US regional banks, halving the modelled NII shock benefit gets you roughly to where the recalibrated number would land. Not perfect, but better than presenting a number you know is wrong.</p>
""",
    },
    {
        "slug": "stop-using-indirect-excel",
        "title": "INDIRECT() in Excel: Why You Should Stop Using It",
        "h1": "INDIRECT() in Excel: Why You Should Stop Using It",
        "keyword": "excel indirect alternative",
        "meta_desc": "INDIRECT() is volatile, breaks silently when sheets are renamed, and can't be traced. Four alternatives by use case for cleaner financial models.",
        "published": "2026-04-21",
        "summary": "INDIRECT() is the most overused finance modelling function. Three specific problems and four alternatives by use case.",
        "product_link": "/models.html",
        "product_name": "SFS Models Catalogue",
        "body": """<p>INDIRECT() is one of those small Excel modelling habits that causes outsized pain when models get inherited. It's convenient for building flexible references &mdash; <code>=INDIRECT("'"&amp;A1&amp;"'!B5")</code> &mdash; but it has three properties that should disqualify it from any model that anyone other than the original builder will use.</p>

<h2>1. It's volatile</h2>

<p>Every recalculation re-evaluates every INDIRECT in the workbook. On a 50-tab model with 200 INDIRECT calls, this is the difference between a 0.5s recalc and a 30s recalc. Volatile functions also trigger recalculation when totally unrelated cells change, which makes audits painful.</p>

<h2>2. It breaks silently when sheets are renamed</h2>

<p>No #REF! error. Just wrong numbers. The user has no idea anything went wrong &mdash; the formula still "works", it just points at nothing now and returns 0 or #REF! that's lost in a sea of other values.</p>

<h2>3. It can't be traced</h2>

<p>F2 doesn't show you what cell INDIRECT is pointing at. F5 &rarr; Special &rarr; Precedents doesn't follow it. You have to mentally evaluate the string concatenation to figure out where the reference goes. In a model audit, this is hours of additional work per page.</p>

<h2>What to use instead</h2>

<p><strong>For cross-sheet references where the sheet is fixed:</strong></p>
<p>Just use a normal reference. <code>='Lending'!B5</code>. No INDIRECT needed.</p>

<p><strong>For lookups across multiple sheets where the sheet name varies:</strong></p>
<p>Use a single consolidation tab that pulls all sheets via fixed references, then INDEX/MATCH against the consolidation. One layer of indirection, fully traceable.</p>

<p><strong>For "sometimes show this, sometimes show that":</strong></p>
<p>CHOOSE or INDEX with a switch cell. Both traceable. Both non-volatile.</p>

<pre><code>=CHOOSE(MATCH(scenario_cell, {"Base","Upside","Downside"}, 0),
        Base_value, Upside_value, Downside_value)</code></pre>

<p><strong>If you genuinely need a dynamic sheet reference (rare):</strong></p>
<p>Build a helper column with the actual values and INDEX/MATCH against that. Avoid INDIRECT.</p>

<h2>The general principle</h2>

<p>In a model that other people will inherit, traceability beats convenience every time. Anything that breaks F5 &rarr; Special &rarr; Precedents is a future bug factory.</p>

<p>For our entire <a href="/models.html">model catalogue</a>, we have a hard rule: no INDIRECT, no INDIRECT.EXT, no OFFSET (similarly volatile and untraceable). Every reference is either direct or via INDEX/MATCH against a fixed range.</p>

<p>It takes 5% more upfront work and saves 50% of audit time. Always worth it.</p>
""",
    },
    {
        "slug": "ccar-stress-capital-buffer-2026",
        "title": "The CCAR Stress Capital Buffer Trap for the 2026 Cycle",
        "h1": "The CCAR Stress Capital Buffer Trap for the 2026 Cycle",
        "keyword": "stress capital buffer ccar",
        "meta_desc": "The 2026 CCAR cycle introduces revised SCB rules that make dividends a binding capital input. Most regional banks haven't updated their models. What changes and how to fix it.",
        "published": "2026-04-22",
        "summary": "Revised SCB rules for the 2026 CCAR cycle make planned dividends a binding input to capital adequacy. Most regional bank models still treat dividends as a pure use of capital. Here's what changes and what to fix.",
        "product_link": "/bank-stress-test-model.html",
        "product_name": "Bank Stress Test Model",
        "body": """<p>The 2026 CCAR cycle is the first under the revised stress capital buffer (SCB) rules. Most US regional banks are still running their 2024 framework. The gap between the two will show up at the next examination cycle.</p>

<h2>What is the Stress Capital Buffer?</h2>

<p>The SCB is the bank-specific buffer above the regulatory minimum CET1 ratio (4.5%) plus the global capital conservation buffer (2.5%). For US banks subject to CCAR, the SCB is calibrated based on the bank's projected capital decline under the supervisory severely adverse scenario.</p>

<p>SCB = Maximum projected capital decline + 4 quarters of planned common stock dividends.</p>

<p>That last component &mdash; planned dividends &mdash; is what changed for 2026.</p>

<h2>What changed for 2026</h2>

<p>Under the previous methodology, planned dividends in the SCB calculation were measured as a static four-quarter sum at submission time. Under the revised methodology, the four-quarter dividend sum is calculated dynamically across the stress horizon.</p>

<p>The practical effect: as your projected dividends change quarter-by-quarter under stress, your SCB recalibrates. Higher dividends = larger SCB = higher minimum CET1 you must maintain.</p>

<p>This makes dividends a <strong>binding input to capital adequacy</strong>, not just a use of capital. If your model treats dividends as a flow to be subtracted from CET1 each quarter (the standard approach), you're missing the buffer feedback loop.</p>

<h2>What breaks in most internal models</h2>

<p>Three patterns we see consistently when reviewing regional bank stress test models:</p>

<ol>
<li><strong>Dividends modelled as a pure use of capital.</strong> Each quarter, dividends reduce CET1. The SCB is a separate input on the capital ratio tab. The two don't talk to each other.</li>
<li><strong>SCB modelled as a single static add-on.</strong> The buffer is a number entered on INPUTS and held constant across the stress horizon. Quarter-by-quarter recalculation isn't built in.</li>
<li><strong>PPNR projections still using 2018-vintage models.</strong> Pre-COVID rate environments don't tell you anything useful about 2024-26. Net interest margin under stress in the new regime depends heavily on deposit beta calibration (see <a href="/blog/deposit-beta-calibration.html">our piece on deposit beta</a>).</li>
</ol>

<h2>The fix: quarterly SCB recalculation</h2>

<p>The model structure that handles the new methodology:</p>

<ul>
<li><strong>One INPUTS tab</strong> with quarterly dividend assumptions (not annual aggregate)</li>
<li><strong>One CAPITAL_ACTIONS tab</strong> calculating the rolling 4-quarter dividend sum at each quarter</li>
<li><strong>One SCB tab</strong> calculating the SCB at each quarter as: max projected capital decline (across remaining horizon from this quarter) + rolling 4-quarter dividend sum</li>
<li><strong>One CAPITAL_RATIOS tab</strong> calculating CET1 ratio against the time-varying SCB-inclusive minimum</li>
<li><strong>MDA logic</strong> that auto-restricts dividends when the CET1 ratio falls below P1 + buffers, which then feeds back into the SCB calculation in the next iteration</li>
</ul>

<p>The feedback loop between dividends and SCB is what makes the new methodology operationally complex. Done right, the model converges quickly. Done wrong (using iterative calculation), it spirals.</p>

<p>Our trick: model the MDA restriction as a function of <em>prior quarter's</em> capital ratio, not current. This breaks the circularity cleanly. Same principle as our universal "no circularities" rule for bank models.</p>

<h2>Walked example: $50bn regional bank</h2>

<p>Consider a $50bn regional bank planning $1.2bn in annual dividends ($300m/quarter).</p>

<p><strong>Under old methodology:</strong> SCB calculation includes $1.2bn dividend addition. Maximum projected CET1 decline (say 250bps over 9 quarters) drives a buffer of, say, 3.5%.</p>

<p><strong>Under new methodology:</strong> at each quarter, the rolling 4-quarter dividend sum is recalculated. If projected stress causes dividend cuts (per MDA), the SCB declines &mdash; but lagging the actual decline. The bank may face a higher SCB during the recovery phase than under the old methodology.</p>

<p>Practical impact: the bank's modelled CET1 decline may now breach the 5.125% AT1 trigger floor when it didn't under the old framework. That's a material capital plan change that needs to land before the 2026 cycle, not during it.</p>

<h2>What to do now</h2>

<ol>
<li><strong>Audit your current capital plan model.</strong> Search for "SCB" in formulas. If it appears as a static cell on INPUTS, you're on the old methodology.</li>
<li><strong>Rebuild capital_actions and SCB calculations as quarterly time series.</strong> Don't shortcut to annual.</li>
<li><strong>Recalibrate PPNR.</strong> Use 2022&ndash;2025 actuals as the base, not 2018&ndash;2019.</li>
<li><strong>Recalibrate deposit beta.</strong> The 2022&ndash;23 cycle showed actual betas were 2&ndash;3x your old assumptions. See <a href="/blog/deposit-beta-calibration.html">deposit beta calibration</a>.</li>
<li><strong>Run the model end-to-end on the 2025 supervisory severely adverse scenario.</strong> If your output looks materially worse than your 2024 submission, you're calibrated correctly. If it looks the same, something's wrong.</li>
</ol>

<h2>The model that handles this</h2>

<p>Our <a href="/bank-stress-test-model.html">Bank Stress Test Model</a> is built for the new methodology natively &mdash; quarterly SCB recalculation, dividend-as-input, recalibratable PPNR, MDA feedback loop. Tested against the 2025 supervisory severely adverse scenario.</p>
""",
    },
    {
        "slug": "cecl-methodology-drift",
        "title": "Why Most CECL Models Drift After Year 2 (And How to Fix It)",
        "h1": "Why Most CECL Models Drift After Year 2 (And How to Fix It)",
        "keyword": "cecl methodology drift",
        "meta_desc": "CECL went live for SEC filers in 2020. Five years later, most US bank methodologies have quietly drifted away from regulatory expectations. The four drift patterns examiners are flagging and how to fix each.",
        "published": "2026-04-22",
        "summary": "CECL adoption was 2020. Five years on, methodologies have drifted. Examiners are flagging four specific patterns: stale R&S forecast horizons, undefended reversion paths, ungrounded Q-factors, and outdated pool segmentation. How to fix each.",
        "product_link": "/ifrs9-cecl-model-excel.html",
        "product_name": "IFRS 9 / CECL ECL Model",
        "body": """<p>CECL (ASC 326) went live for SEC filers in 2020 and for non-SEC filers in 2023. Most banks built their methodology in the adoption window and haven't materially refreshed it since. That's a problem &mdash; OCC and FDIC examiners are now flagging methodology drift as a systemic issue, particularly at regional banks with CRE concentration.</p>

<h2>The four drift patterns</h2>

<h3>1. Reasonable & Supportable forecast horizon hasn't been re-justified</h3>

<p>CECL requires a Reasonable &amp; Supportable (R&amp;S) forecast period &mdash; the window over which you can defensibly project macro variables. After R&amp;S, you revert to historical loss rates.</p>

<p>Most banks picked 1&ndash;2 years at adoption and haven't revisited it. In 2020, with COVID uncertainty, 1 year was defensible. In 2026, with a clearer macro outlook, you might justify 2&ndash;3 years &mdash; but you must document why.</p>

<p><strong>Examiner flag:</strong> "Why is your R&amp;S period still 1 year? What changed since adoption to justify the same window?" If you don't have a written answer, you have a finding.</p>

<p><strong>Fix:</strong> annual R&amp;S period review, board-approved. Document the macro backdrop justification each year.</p>

<h3>2. Reversion path is an Excel formula nobody can defend</h3>

<p>After the R&amp;S period, you must revert from forecast loss rates to historical mean. Most banks implemented this as a linear interpolation in Excel and never wrote the methodology document.</p>

<p>The reversion path matters: a steep reversion in year 2 vs a gentle reversion over years 2&ndash;5 produces materially different lifetime ECL.</p>

<p><strong>Examiner flag:</strong> "Show me the methodology document for your reversion path. Why linear vs exponential? Why over this duration?"</p>

<p><strong>Fix:</strong> document the reversion methodology explicitly. Common defensible approaches:</p>
<ul>
<li>Linear over R&amp;S+1 to year 5 (simple, defensible if no specific macro view)</li>
<li>Exponential decay (industry-standard for some portfolios)</li>
<li>Step-function aligned to forecast cycles (e.g., revert at end of forecast)</li>
</ul>

<h3>3. Q-factor adjustments don't tie to specific defensible signals</h3>

<p>The Q-factor (qualitative adjustment) is where most banks bury the macro overlay. The amount is defensible only if it ties to a specific external signal.</p>

<p>"Macro deterioration" isn't a number. "+10bps to allowance reflecting CRE office vacancy increase from 14% to 19% per CBRE Q4 2025 report" is a number.</p>

<p><strong>Examiner flag:</strong> "What's the source for your +5bps Q-factor on the C&amp;I book?" If you can't cite an external data source, the Q-factor will be challenged.</p>

<p><strong>Fix:</strong> every Q-factor entry must have:</p>
<ul>
<li>An external data source (Fed, Bloomberg, CBRE, BLS, etc.)</li>
<li>A documented rule connecting the data signal to the Q-factor magnitude</li>
<li>Evidence of historical back-testing (does the rule produce sensible numbers in past cycles?)</li>
</ul>

<h3>4. Pool segmentation hasn't evolved with portfolio mix</h3>

<p>Your 2020 CRE pool segmentation (owner-occupied / income-producing / construction) may have been adequate then. In 2026, with the office sector deterioration creating a wholly different risk profile, you might need to subdivide income-producing into office / multifamily / retail / industrial.</p>

<p><strong>Examiner flag:</strong> "Why is your office CRE pooled with multifamily?" If you don't have a defensible reason, that's a finding.</p>

<p><strong>Fix:</strong> annual segmentation review. If a sub-segment has materially different risk characteristics from the parent pool, separate it.</p>

<h2>The annual methodology refresh framework</h2>

<p>Build this into your CECL governance:</p>

<ol>
<li><strong>Annual R&amp;S period review</strong> &mdash; 1 page, board-approved, justifies the chosen horizon</li>
<li><strong>Annual reversion methodology review</strong> &mdash; 1 page, justifies the chosen reversion shape and duration</li>
<li><strong>Quarterly Q-factor review</strong> &mdash; line-by-line update of each Q-factor with external data citation</li>
<li><strong>Annual pool segmentation review</strong> &mdash; assesses whether the current pools are still risk-coherent</li>
<li><strong>Annual back-test</strong> &mdash; compares the prior year's allowance against actual losses, identifies methodology gaps</li>
</ol>

<p>This isn't optional. Examiners explicitly ask for the documentation. If it doesn't exist, you have a finding.</p>

<h2>The model that supports this</h2>

<p>Our <a href="/ifrs9-cecl-model-excel.html">IFRS 9 / CECL ECL Model</a> builds in the documentation framework: Q-factor cells include source citation columns, R&amp;S and reversion parameters live in a single methodology tab, and pool segmentation is rebuildable without breaking the calculation layer.</p>

<p>Audit-ready by construction. Methodology document is the same workbook you use to compute the allowance.</p>
""",
    },
    {
        "slug": "python-openpyxl-financial-model",
        "title": "Building Financial Models Programmatically with Python and openpyxl",
        "h1": "Building Financial Models Programmatically with Python and openpyxl",
        "keyword": "openpyxl financial model",
        "meta_desc": "How we build 18 institutional-grade Excel financial models programmatically using Python and openpyxl. Patterns, gotchas, and the standard helper library.",
        "published": "2026-04-22",
        "summary": "Hand-edited Excel models are unmaintainable at scale. Python + openpyxl makes the build script the source of truth. Here's our standard pattern: helpers, named ranges, the no-circularity rule, and how we validate all 18 models in one command.",
        "product_link": "/models.html",
        "product_name": "SFS Models Catalogue",
        "body": """<p>Every financial model in our catalogue is generated by a Python script using openpyxl. We never edit the .xlsx files by hand. The build script is the source of truth.</p>

<p>This isn't a hipster tooling choice &mdash; it's the only way to maintain a catalogue of 18 institutional-grade models without going insane. Here's the pattern.</p>

<h2>Why scripts as source of truth</h2>

<p>Hand-edited Excel files have four problems at scale:</p>

<ol>
<li><strong>No version control</strong>. Git can store .xlsx but can't diff it. You can see <em>that</em> the file changed but not <em>what</em> changed.</li>
<li><strong>Style drift</strong>. Five models, five slightly different formats. Looks unprofessional.</li>
<li><strong>Cascading bugs</strong>. Fix a formula in one model, forget to fix it in the other 17.</li>
<li><strong>No testing</strong>. You can't unit-test a hand-edited Excel file. You can unit-test a Python script that generates one.</li>
</ol>

<p>Build scripts solve all four. Each model has a build_<em>name</em>_v<em>n</em>.py file. Run the script, get the .xlsx. Change the script, regenerate. Commit the script to git.</p>

<h2>The standard helper library</h2>

<p>Every build script imports the same set of helpers:</p>

<pre><code class="language-python">from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.workbook.defined_name import DefinedName

# Style constants
NAVY       = "1F3864"
MID_BLUE   = "2E75B6"
WHITE      = "FFFFFF"
BLUE_FONT  = "0000FF"   # hardcoded inputs
BLACK_FONT = "000000"   # formulas
YELLOW_BG  = "FFFDE7"   # input cell fill
GREEN_PASS = "C6EFCE"   # CHECKS pass
RED_FAIL   = "FFC7CE"   # CHECKS fail
GREEN_TOT  = "E2EFDA"   # total row fill

# Number formats
FMT_GBP    = '#,##0;(#,##0);"-"'
FMT_GBPm   = '#,##0.0,,"m";(#,##0.0,,"m");"-"'
FMT_PCT    = '0.0%;(0.0%);"-"'
FMT_BPS    = '0" bps";(0" bps");"-"'
FMT_DATE   = 'mmm-yy'

def inp(ws, row, col, value, fmt=FMT_GBP):
    # Hardcoded input cell: blue font, yellow fill.
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Arial", size=10, color=BLUE_FONT)
    c.fill = PatternFill("solid", fgColor=YELLOW_BG)
    c.number_format = fmt
    return c

def calc(ws, row, col, formula, fmt=FMT_GBP):
    # Formula cell: black font, no fill.
    c = ws.cell(row=row, column=col, value=formula)
    c.font = Font(name="Arial", size=10, color=BLACK_FONT)
    c.number_format = fmt
    return c

def label(ws, row, col, text, indent=0, bold=False):
    # Text label cell.
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(name="Arial", size=10, bold=bold)
    c.alignment = Alignment(horizontal="left", indent=indent)
    return c

def section_header(ws, row, col_start, col_end, text):
    # Dark navy section header.
    c = ws.cell(row=row, column=col_start, value=text)
    c.font = Font(name="Arial", size=10, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=NAVY)
    for col in range(col_start + 1, col_end + 1):
        ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=NAVY)
    return c
</code></pre>

<p>Three functions handle 80% of cell creation: <code>inp()</code> for hardcoded inputs (blue/yellow), <code>calc()</code> for formulas (black, no fill), <code>label()</code> for text. Consistency across all 18 models comes for free.</p>

<h2>Common openpyxl gotchas</h2>

<h3>Named ranges API changed</h3>

<p>The old <code>wb.defined_names.definedName</code> API is deprecated. Use string-key iteration:</p>

<pre><code class="language-python"># CORRECT
for name in wb.defined_names:
    dn = wb.defined_names[name]

# WRONG (raises AttributeError in current openpyxl)
for dn in wb.defined_names.definedName:
    ...
</code></pre>

<h3>Formulas are strings, not evaluated</h3>

<p>openpyxl writes formula strings to cells but doesn't evaluate them. The cell's value when read back is the formula text, not the result. To get evaluated values, you need to either:</p>

<ul>
<li>Open the file in Excel/LibreOffice (which triggers recalc on save)</li>
<li>Run the file through <code>libreoffice --headless --convert-to xlsx</code> to force recalc</li>
<li>Use a separate library like <code>pycel</code> (limited support) or <code>formulas</code> (better but slow)</li>
</ul>

<p>Our validation pipeline runs every model through headless LibreOffice to force recalc, then reads the values back. Catches formula bugs before shipping.</p>

<h3>Performance with large workbooks</h3>

<p>openpyxl in default (cell-by-cell) mode is slow for large workbooks. For models with &gt;10,000 formula cells, use write-only mode or batch the cell creation.</p>

<p>Bigger payoff: avoid creating empty cells. openpyxl tracks every accessed cell, so even reading <code>ws.cell(row=1000, col=1)</code> in an empty area inflates the file. Only touch cells you intend to populate.</p>

<h2>The no-circularity rule</h2>

<p>The hardest discipline in financial modelling: avoiding circular references.</p>

<p>The classic case in banking: interest income depends on the loan balance, but the loan balance depends on interest income (interest gets capitalised, or the cash book grows from interest received).</p>

<p>The naive fix &mdash; turn on Excel's iterative calculation &mdash; works but creates a brittle model. Audit teams reject models that depend on iterative calculation because they can produce different results based on iteration count.</p>

<p>The clean fix: <strong>always reference prior period balance, never current</strong>.</p>

<pre><code class="language-python"># Cell layout: opening, drawdown, repayment, interest, closing
# For row 12 (period 2):
calc(ws, 12, 5, '=I11')              # opening = prior closing
calc(ws, 12, 8, '=E12*rate/12')      # interest on opening
calc(ws, 12, 9, '=E12+F12-G12+H12')  # closing = opening + flows + interest
</code></pre>

<p>This pattern works for: lending interest, deposit interest, capital constraints (current period capital from prior period), provision charges. Circular by reference, sequential by construction.</p>

<h2>Validation: the CHECKS tab</h2>

<p>Every model has a CHECKS tab with 30&ndash;40 integrity tests:</p>

<ul>
<li>BS balances (assets = liabilities + equity)</li>
<li>CF closing ties to BS cash</li>
<li>Debt schedule closing ties to BS debt</li>
<li>Total interest from debt schedule = total interest in IS</li>
<li>Scenario switches working (Upside &gt; Base on key metrics)</li>
<li>No negative equity in Base scenario</li>
</ul>

<p>Each check returns "PASS" or "FAIL" with conditional formatting (green/red). The whole catalogue is validated by:</p>

<pre><code class="language-bash">cd models &amp;&amp; python3 validate_all.py</code></pre>

<p>If any model has a failing CHECK, the build fails. Catches bugs before shipping.</p>

<h2>The full pipeline</h2>

<ol>
<li>Edit build_<em>name</em>_v<em>n</em>.py</li>
<li>Run <code>python3 build_<em>name</em>_v<em>n</em>.py</code> &mdash; outputs .xlsx</li>
<li>Run <code>python3 validate_all.py</code> &mdash; opens each .xlsx through LibreOffice headless and reads CHECKS tab</li>
<li>If validation passes, commit the build script + .xlsx to git</li>
<li>Push &mdash; CI rebuilds and re-validates as a sanity check</li>
</ol>

<p>This is how we ship 18 institutional-grade models with two engineers (effectively just me) and zero production bugs in 18 months.</p>

<h2>Should you do this?</h2>

<p>If you're building 1&ndash;2 financial models, no &mdash; openpyxl is overkill, just use Excel directly.</p>

<p>If you're building 5+ models that share styling, structure, or reusable calculations, yes &mdash; the upfront investment in helpers pays back within the third model.</p>

<p>Our entire <a href="/models.html">catalogue of 18 models</a> uses this approach. If you're curious about the helper library or have specific openpyxl questions, ping us via <a href="/contact.html">contact</a>.</p>
""",
    },
]


PAGE_TEMPLATE = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>{title} | SFS Models Blog</title>
<meta name="description" content="{meta_desc}">
<link rel="canonical" href="https://sfsmodels.org/blog/{slug}.html">

<meta property="og:type" content="article">
<meta property="og:title" content="{title}">
<meta property="og:description" content="{meta_desc}">
<meta property="og:url" content="https://sfsmodels.org/blog/{slug}.html">
<meta property="og:site_name" content="SFS Models">

<script type="application/ld+json">
{{
  "@context": "https://schema.org",
  "@type": "BlogPosting",
  "headline": "{title}",
  "description": "{meta_desc}",
  "datePublished": "{published}",
  "dateModified": "{published}",
  "author": {{ "@type": "Organization", "name": "SFS Models", "url": "https://sfsmodels.org" }},
  "publisher": {{ "@type": "Organization", "name": "SFS Models", "url": "https://sfsmodels.org" }},
  "mainEntityOfPage": "https://sfsmodels.org/blog/{slug}.html"
}}
</script>

<link rel="icon" href="data:image/svg+xml,<svg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 32 32'><rect width='32' height='32' rx='6' fill='%23C9A84C'/><text x='16' y='23' font-family='system-ui' font-size='20' font-weight='700' fill='%230A0C10' text-anchor='middle'>S</text></svg>">
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=DM+Sans:wght@400;500;600;700&family=DM+Serif+Display&display=swap" rel="stylesheet">
<link rel="stylesheet" href="../css/style.css">
<style>
.blog-post {{ max-width: 720px; margin: 0 auto; padding: 2rem 0; }}
.blog-post h1 {{ font-family: 'DM Serif Display', serif; font-size: 2.4rem; line-height: 1.15; margin-bottom: 0.5rem; }}
.blog-post .meta {{ color: var(--text-muted); font-size: 0.9rem; margin-bottom: 2rem; }}
.blog-post .summary {{ font-size: 1.15rem; line-height: 1.55; padding: 1rem 1.25rem; border-left: 3px solid #C9A84C; background: rgba(201,168,76,0.05); margin-bottom: 2rem; }}
.blog-post h2 {{ font-family: 'DM Serif Display', serif; font-size: 1.55rem; margin-top: 2.2rem; margin-bottom: 0.6rem; }}
.blog-post p, .blog-post li {{ line-height: 1.7; font-size: 1.02rem; }}
.blog-post pre {{ background: rgba(0,0,0,0.06); padding: 0.9rem 1rem; border-radius: 6px; overflow-x: auto; font-size: 0.92rem; }}
.blog-post code {{ font-family: ui-monospace, 'SF Mono', Menlo, Consolas, monospace; }}
.blog-post a {{ color: #C9A84C; }}
.blog-cta {{ margin: 3rem 0; padding: 1.5rem; border: 1px solid #C9A84C; border-radius: 8px; }}
.blog-cta h3 {{ margin-top: 0; font-family: 'DM Serif Display', serif; }}
</style>
</head>
<body>

<a href="#main-content" class="skip-link">Skip to main content</a>

<header class="site-header">
  <div class="container">
    <nav class="nav-inner">
      <a href="/index.html" class="logo">SFS Models</a>
      <ul class="nav-links">
        <li><a href="/models.html">Models</a></li>
        <li><a href="/previews.html">Preview Models</a></li>
        <li><a href="/free-samples.html">Free Samples</a></li>
        <li><a href="/about.html">About</a></li>
        <li><a href="/contact.html">Contact</a></li>
      </ul>
      <div class="nav-cta">
        <a href="/contact.html" class="btn btn-primary btn-sm">Get a Custom Model</a>
      </div>
      <button class="hamburger" aria-label="Menu" aria-expanded="false"><span></span><span></span><span></span></button>
    </nav>
  </div>
</header>

<main id="main-content">
  <article class="container blog-post">
    <p style="color:var(--text-muted); font-size:0.85rem; margin-bottom: 0.5rem;"><a href="/" style="color:var(--text-muted);">Home</a> &middot; <a href="/blog/" style="color:var(--text-muted);">Blog</a></p>
    <h1>{h1}</h1>
    <p class="meta">Published {published} &middot; SFS Models</p>
    <p class="summary">{summary}</p>

    {body}

    <div class="blog-cta">
      <h3>Get the {product_name}</h3>
      <p>The model that puts the principles in this post into practice. Bank-grade build, open formulas, no VBA. Same-day delivery.</p>
      <p><a href="{product_link}" class="btn btn-primary">View {product_name}</a> &nbsp; <a href="/free-sample.html" class="btn btn-secondary">Try Free Sample</a></p>
    </div>

    <p style="margin-top:3rem; padding-top:1.5rem; border-top:1px solid rgba(0,0,0,0.1); font-size:0.95rem;">SFS Models builds institutional-grade Excel financial models for banking and finance professionals. <a href="/models.html">Browse the catalogue</a> or get the <a href="/free-sample.html">free sample</a>.</p>
  </article>
</main>

<footer class="site-footer">
  <div class="container">
    <div class="footer-grid">
      <div class="footer-brand">
        <a href="/index.html" class="logo">SFS Models</a>
        <p>Institutional-grade financial models for banks, fintechs, and advisory firms. Pre-built or custom.</p>
      </div>
      <div class="footer-col">
        <h4>Models</h4>
        <ul>
          <li><a href="/models.html#banking-lending">Banking &amp; Lending</a></li>
          <li><a href="/models.html#treasury-capital-markets">Treasury</a></li>
          <li><a href="/models.html#risk-regulatory">Risk &amp; Regulatory</a></li>
          <li><a href="/models.html#valuation-corporate-finance">Valuation</a></li>
          <li><a href="/models.html#fpa-management-reporting">FP&amp;A</a></li>
        </ul>
      </div>
      <div class="footer-col">
        <h4>Company</h4>
        <ul>
          <li><a href="/about.html">About</a></li>
          <li><a href="/about.html#methodology">Methodology</a></li>
          <li><a href="/pricing.html#faq">FAQ</a></li>
          <li><a href="/contact.html">Contact</a></li>
        </ul>
      </div>
    </div>
    <div class="footer-bottom">
      <span>&copy; 2026 SFS Models. All rights reserved.</span>
      <span><a href="/terms.html">Terms &amp; Conditions</a> &middot; <a href="/privacy.html">Privacy Policy</a> &middot; London, UK &middot; sfsmodels362@gmail.com</span>
    </div>
  </div>
</footer>

<script src="/js/scripts.js"></script>
</body>
</html>
"""


def main():
    out_dir = Path(__file__).parent / "blog"
    out_dir.mkdir(exist_ok=True)
    for p in POSTS:
        html = PAGE_TEMPLATE.format(**p)
        out_path = out_dir / f"{p['slug']}.html"
        out_path.write_text(html, encoding="utf-8")
        print(f"Wrote blog/{p['slug']}.html ({len(html):,} bytes)")

    # Build a simple /blog/index.html
    items = []
    for p in POSTS:
        items.append(
            f'<li style="margin-bottom:1.5rem;"><a href="/blog/{p["slug"]}.html" style="color:#C9A84C; font-size:1.15rem; font-weight:600;">{p["title"]}</a><br><span style="color:var(--text-muted); font-size:0.95rem;">{p["summary"]}</span><br><span style="color:var(--text-muted); font-size:0.85rem;">Published {p["published"]}</span></li>'
        )
    index_body = "\n".join(items)
    index_html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Blog | SFS Models</title>
<meta name="description" content="Banking, finance, and modelling articles from SFS Models. Bank-grade Excel financial models for finance professionals.">
<link rel="canonical" href="https://sfsmodels.org/blog/">
<link rel="icon" href="data:image/svg+xml,<svg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 32 32'><rect width='32' height='32' rx='6' fill='%23C9A84C'/><text x='16' y='23' font-family='system-ui' font-size='20' font-weight='700' fill='%230A0C10' text-anchor='middle'>S</text></svg>">
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=DM+Sans:wght@400;500;600;700&family=DM+Serif+Display&display=swap" rel="stylesheet">
<link rel="stylesheet" href="../css/style.css">
</head>
<body>
<a href="#main-content" class="skip-link">Skip to main content</a>
<header class="site-header"><div class="container"><nav class="nav-inner"><a href="/index.html" class="logo">SFS Models</a><ul class="nav-links"><li><a href="/models.html">Models</a></li><li><a href="/previews.html">Preview Models</a></li><li><a href="/free-samples.html">Free Samples</a></li><li><a href="/about.html">About</a></li><li><a href="/contact.html">Contact</a></li></ul></nav></div></header>
<main id="main-content">
<section class="page-hero"><div class="container"><h1>Blog</h1><p>Banking, finance, and modelling articles. New posts weekly.</p></div></section>
<section class="section"><div class="container" style="max-width:760px;"><ul style="list-style:none; padding:0;">
{index_body}
</ul></div></section>
</main>
<footer class="site-footer"><div class="container"><div class="footer-bottom"><span>&copy; 2026 SFS Models. All rights reserved.</span><span><a href="/terms.html">Terms</a> &middot; <a href="/privacy.html">Privacy</a> &middot; London, UK &middot; sfsmodels362@gmail.com</span></div></div></footer>
</body>
</html>
"""
    (out_dir / "index.html").write_text(index_html, encoding="utf-8")
    print(f"Wrote blog/index.html ({len(index_html):,} bytes)")


if __name__ == "__main__":
    main()
